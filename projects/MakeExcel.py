import os
import openpyxl

os.chdir(r"C:\Users\safab\OneDrive\Desktop")

workbook = openpyxl.Workbook() 
sheet1 = workbook.active # Look for the active sheet

sheet1.cell(row=1, column=1).value = "Names"
sheet1.cell(row=1, column=2).value = "Ages"

workbook.save("secret1.xlsx")
