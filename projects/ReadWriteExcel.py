import os
import openpyxl

os.chdir(r"C:\Users\safab\OneDrive\Desktop")

workbook = openpyxl.load_workbook("secret.xlsx")
sheet1 = workbook["Sheet1"] # Go to sheet 1
print(sheet1["A3"].value) # Go to the cell A3 and for .value to gives as just the valur of the cell
sheet1["A6"].value = "Safae" #The file .xlsx should not be opened
sheet1["B6"].value = "20"
workbook.save("secret.xlsx")

print(sheet1.cell(row=1, column=1).value)

for i in range(1,7):
    print(sheet1.cell(row=i, column=1).value , sheet1.cell(row=i, column=2).value)
